import requests
import xml.etree.ElementTree as ET
import openpyxl

class ExcelWriter():
    def write_data(self, list):
        path = "C:/Users/Acer/Documents/GitHub/Wrblr/chat app 2/online_dataimporter"
        file = "Excel.xlsx"
        wb = openpyxl.load_workbook(str(path) +'/' + str(file))
        sheet = wb.active

        headers = ['COMMON', 'BOTANICAL', 'ZONE', 'LIGHT', 'PRICE', 'AVAILABILITY']
        col = headers.index(list[0]) + 1

        empty_cell = 0
        for line in list:
            empty_cell = empty_cell+1#sheet.max_row + 1
            if line != "":
              sheet.cell(row=empty_cell, column=col).value = line
            wb.save(path+"/"+file)
            #print(str(line)+"saved to" +str(empty_cell))

class APICaller():
    def make_api_call(self, link):
        #params ={"valuta": "eur"}
        #headers = {'Accept': 'application/xml'}
        response = requests.get(link)#, params=params, headers=headers)
        if response.status_code == 200:
            xml_content = response.content
            return(xml_content)

class IOHandler():
    def read_input(self):
        result = input()
        return result
    
    def write_output(self, output):
        print(output)

class ListFromXMLData():
    def CreateList(self, xml, TAG):
        root = ET.fromstring(xml)
        selected_data=[TAG]
        for plant in root.findall("PLANT"):
            try:
                data_row = plant.find(TAG).text
            except:
                data_row = ""
            
            selected_data.append(data_row)
        return(selected_data)
    
#child.text for child in element
#element.tag
#ezek nehezen implementálhatóak a jelen XML fájlra, mert annak felépítése az eredeti feladatnál lényegesen egyszerübb