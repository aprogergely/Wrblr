import pytest
from webscraper import APICaller
from webscraper import ListFromXMLData
from webscraper import ExcelWriterTestable
import openpyxl
import os

def testAPICaller():
    arfolyam = APICaller()
    try:
      flowers=arfolyam.make_api_call("http://api.napiarfolyam.hu/?bank=kh&valutanem=valuta")
      assert True
    except:
      assert False



def APICallerTest():
   flowers ='''
   <arfolyamok>
      <valuta>
          <item>
            <bank>kh</bank>
            <datum>2023-06-29 06:33:05</datum>
            <penznem>GBP</penznem>
            <vetel>415.6700</vetel>
            <eladas>445.8300</eladas>
          </item>
          <item>
            <bank>kh</bank>
            <datum>2023-06-29 06:33:05</datum>
            <penznem>AUD</penznem>
            <vetel>215.3700</vetel>
            <eladas>235.6700</eladas>
          </item>
      </valuta>
    </arfolyamok>
    '''
   return flowers



def testListFromXMLData():
    arfolyam = APICaller()
    flowers=arfolyam.make_api_call("http://api.napiarfolyam.hu/?bank=kh&valutanem=valuta")

    xmldata = ListFromXMLData()
    list=xmldata.CreateList(flowers, "penznem")
    expected_output = "AUD"

    actual_output = list[2]

    assert actual_output == expected_output




def testListFromXMLData2():
    flowers=APICallerTest()
    xmldata = ListFromXMLData()
    list=xmldata.CreateList(flowers, "penznem")
    expected_output = "AUD"

    actual_output = list[2]

    assert actual_output == expected_output




@pytest.mark.parametrize('path' , ['C:/Users/Acer/Documents/GitHub/Wrblr/chat app 2/online_dataimporter'])
@pytest.mark.parametrize('backup_file' , ['Excel_backup.xlsx'])
@pytest.mark.parametrize('file' , ['Excel.xlsx'])
def testExcelWriter(path, file, backup_file):
    wb = openpyxl.load_workbook(str(path) +'/' + str(file))
    wb.save(str(path) +'/' + str(backup_file))
    wb.close()

    excel=ExcelWriterTestable()
    list=("vetel", "hello", "pingvin", "43edg52")
    excel.write_data(list, path, backup_file)

    wb = openpyxl.load_workbook(str(path) +'/' + str(backup_file))
    sheet = wb.active

    actual_output=sheet.cell(row=2, column=4).value
    expected_output = "hello"

    os.remove(str(path) +'/' + str(backup_file))
    assert actual_output == expected_output