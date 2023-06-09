import requests
import xml.etree.ElementTree as ET
import openpyxl
import os

class ExcelWriter():
    def write_data(self, list):
        script_directory = os.path.dirname(os.path.realpath(__file__))
        config_file_path = os.path.join(script_directory, 'Config.txt')

        with open(config_file_path, 'r') as config_file:
            for line in config_file:
                if line.startswith('save_location'):
                    save_location = line.split('=')[1].strip()
                break

        path = os.path.dirname(save_location)
        file = "Excel.xlsx"
        wb = openpyxl.load_workbook(str(path) +'/' + str(file))
        sheet = wb.active

        headers = ['bank', 'datum', 'penznem', 'vetel', 'eladas']
        col = headers.index(list[0]) + 1

        empty_cell = 0
        for line in list:
            empty_cell = empty_cell+1#sheet.max_row + 1
            if line != "":
              sheet.cell(row=empty_cell, column=col).value = line
            wb.save(path+"/"+file)
            #print(str(line)+"saved to" +str(empty_cell))

class APICaller():
    def make_api_call(self, link):
        #params ={"valuta": "eur"}
        #headers = {'Accept': 'application/xml'}
        response = requests.get(link)#, params=params, headers=headers)
        if response.status_code == 200:
            xml_content = response.content
            return(xml_content)

class IOHandler():
    def read_input(self):
        result = input()
        return result
    
    def write_output(self, output):
        print(output)

class ListFromXMLData():
    def CreateList(self, xml, TAG):
        root = ET.fromstring(xml)
        selected_data=[TAG]
        for element in root.findall("valuta"):
            for child in element.findall("item"):
                try:
                    data_row = child.find(TAG).text
                except:
                    data_row = ""
            
                selected_data.append(data_row)
        return(selected_data)
    
#child.text for child in element
#element.tag


class ExcelWriterTestable():
    def write_data(self, list, path, file):
        path = path
        file = file
        wb = openpyxl.load_workbook(str(path) +'/' + str(file))
        sheet = wb.active

        headers = ['bank', 'datum', 'penznem', 'vetel', 'eladas']
        col = headers.index(list[0]) + 1

        empty_cell = 0
        for line in list:
            empty_cell = empty_cell+1#sheet.max_row + 1
            if line != "":
              sheet.cell(row=empty_cell, column=col).value = line
            wb.save(path+"/"+file)
            #print(str(line)+"saved to" +str(empty_cell))